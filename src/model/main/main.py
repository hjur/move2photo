import sys
import openpyxl
from PySide6.QtWidgets import *
from src.UI.main.mainUI import Ui_MainWindow

class mainUI(QMainWindow, Ui_MainWindow):
        def __init__(self, *args, obj=None, **kwargs):
            super(mainUI, self).__init__(*args, **kwargs)
            self.setupUi(self)
            self.initEvent()

        def initEvent(self):
            self.btnOpenExcel.clicked.connect(self.clickOpenExcel)

        def clickOpenExcel(self):
            fd = QFileDialog()
            # 다이얼로그 옵션 설정 및 오픈.
            fname = fd.getOpenFileName(self, "Open File", 'd:\\', '엑셀 파일 (*.xlsx *.xls)')
            print('fname[0]:'+fname[0])
            # 선택된 엑셀 파일을 읽기. data_only=True : 수식도 값으로 읽기
            wb = openpyxl.load_workbook(str(fname[0]), data_only=True)
            # 첫번째 시트만 읽어 온다
            ws = wb[wb.sheetnames[0]]
            # 한줄씩 읽는다 . 파일이 크면 오래걸리겠지.. 개선 필요
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                value_list = []
                for cell in row:
                    value = cell.value if cell.value else ''
                    value_list.append('{:>15}'.format(value))
                print(''.join(value_list))
            wb.close()
            # self.lblExcelName.setText(str(value_list))


app = QApplication()
window = mainUI()
window.show()
app.exec()